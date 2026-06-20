"""报表导出字节构造单元测试（离线、确定性）。"""
import io

from app.reports.export import build_csv, build_xlsx
from openpyxl import load_workbook

HEADERS = ["ID", "级别", "数值"]
ROWS = [[1, "紧急", 3.14], [2, "提示", None]]


def test_build_csv_has_bom_and_rows() -> None:
    data = build_csv(HEADERS, ROWS)
    # UTF-8 BOM 前缀，保证 Excel 中文不乱码
    assert data.startswith(b"\xef\xbb\xbf")
    text = data.decode("utf-8-sig")
    lines = text.splitlines()
    assert lines[0] == "ID,级别,数值"
    assert lines[1] == "1,紧急,3.14"
    # None 输出为空字段
    assert lines[2] == "2,提示,"


def test_build_xlsx_roundtrip() -> None:
    data = build_xlsx(HEADERS, ROWS, "告警")
    wb = load_workbook(io.BytesIO(data))
    ws = wb.active
    assert ws is not None
    assert ws.title == "告警"
    assert [c.value for c in ws[1]] == HEADERS
    assert [c.value for c in ws[2]] == [1, "紧急", 3.14]
    # None 输入写为空单元格（openpyxl 读回为 None）
    assert [c.value for c in ws[3]] == [2, "提示", None]


def test_build_xlsx_long_sheet_name_truncated() -> None:
    data = build_xlsx(HEADERS, ROWS, "a" * 40)
    wb = load_workbook(io.BytesIO(data))
    assert wb.active is not None
    # Excel sheet 名上限 31 字符
    assert len(wb.active.title) == 31
